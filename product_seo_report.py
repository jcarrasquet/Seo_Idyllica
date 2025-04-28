import os
import requests
import pandas as pd
from datetime import date, timedelta
from google.oauth2 import service_account
from googleapiclient.discovery import build
from base64 import b64encode
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import config
import json

# === HEADERS ===
HEADERS = {
    "Authorization": "Basic " + b64encode(f"{config.CONSUMER_KEY}:{config.CONSUMER_SECRET}".encode()).decode("utf-8")
}

# === PRODUCTOS ===
def get_all_products():
    url = config.WC_API_URL
    all_products = []
    page = 1
    while True:
        response = requests.get(url, headers=HEADERS, params={"per_page": 100, "page": page})
        if response.status_code != 200:
            break
        data = response.json()
        if not data:
            break
        all_products.extend(data)
        page += 1
    return all_products

# === Search Console ===
def fetch_seo_data(urls):
    scopes = ["https://www.googleapis.com/auth/webmasters.readonly"]
    credentials_json_str = os.environ["GOOGLE_CREDENTIALS_JSON"]
    credentials_info = json.loads(credentials_json_str)
    credentials = service_account.Credentials.from_service_account_info(credentials_info, scopes=scopes)
    service = build("searchconsole", "v1", credentials=credentials)

    end_date = date.today()
    start_date = end_date - timedelta(days=28)

    seo_rows = []

    for url in urls:
        body = {
            "startDate": str(start_date),
            "endDate": str(end_date),
            "dimensions": ["query"],
            "dimensionFilterGroups": [{
                "filters": [{
                    "dimension": "page",
                    "operator": "equals",
                    "expression": url
                }]
            }],
            "rowLimit": 10
        }
        try:
            response = service.searchanalytics().query(siteUrl="https://www.idyllica.es/", body=body).execute()
            rows = response.get("rows", [])
            keywords = [r["keys"][0] for r in rows]
            position = sum([r.get("position", 0) for r in rows]) / len(rows) if rows else ""
            ctr = sum([r.get("ctr", 0) for r in rows]) / len(rows) if rows else ""
            seo_rows.append((url, keywords, round(position, 2), round(ctr * 100, 2) if ctr else ""))
        except Exception as e:
            seo_rows.append((url, [], "", ""))
    return seo_rows

# === MAIN ===
products = get_all_products()
product_urls = [p["permalink"] for p in products]
seo_data = fetch_seo_data(product_urls)

rows = []
for product, seo in zip(products, seo_data):
    rows.append({
        "Producto": product["name"],
        "URL": product["permalink"],
        "Título SEO": product.get("yoast_head_json", {}).get("title", ""),
        "Meta descripción": product.get("yoast_head_json", {}).get("description", ""),
        "Keywords": ", ".join(seo[1]),
        "Posición media": seo[2],
        "CTR (%)": seo[3]
    })

filename = "seo_report_productos.xlsx"
df = pd.DataFrame(rows)
df.to_excel(filename, index=False)

# === Formato condicional y ajuste de columnas ===
wb = load_workbook(filename)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

# Índice de columna de "Posición media"
position_col_index = None
for idx, cell in enumerate(ws[1], 1):
    if cell.value == "Posición media":
        position_col_index = idx
        break

# Aplicar formato condicional
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    position_cell = row[position_col_index - 1]
    try:
        position = float(position_cell.value)
        if position <= 10:
            for cell in row:
                cell.fill = green_fill
        elif 11 <= position <= 20:
            for cell in row:
                cell.fill = orange_fill
    except:
        continue

# Ajustar anchos
for column_cells in ws.columns:
    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
    adjusted_width = length + 2
    col_letter = column_cells[0].column_letter
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(filename)
print("✅ Informe SEO generado con colores y ajuste de columnas.")

# === EMAIL ===
EMAIL_SENDER = os.getenv("EMAIL_USERNAME")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_RECIPIENT = EMAIL_SENDER

msg = MIMEMultipart()
msg["From"] = EMAIL_SENDER
msg["To"] = EMAIL_RECIPIENT
msg["Subject"] = "📊 Informe SEO semanal - Productos Idyllica"

part = MIMEBase("application", "octet-stream")
with open(filename, "rb") as file:
    part.set_payload(file.read())
encoders.encode_base64(part)
part.add_header("Content-Disposition", f"attachment; filename={filename}")
msg.attach(part)

try:
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, EMAIL_RECIPIENT, msg.as_string())
    print("📧 Email enviado correctamente.")
except Exception as e:
    print(f"❌ Error al enviar el email: {e}")
